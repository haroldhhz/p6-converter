"""
FastAPI backend — P6 Schedule Converter / LAP Platform.

POST /convert/
  Body (multipart/form-data):
    - schedule_file : PDF  — The schedule export from P6
    - standard_file : XLSX — Attachment 3: standard Activity ID + Name + Status
    - data_date     : str  — (optional) IMS date, YYYY-MM-DD

  Returns: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
           (p6_import.xlsx for direct re-import into P6)

Authentication:
  Microsoft Entra ID SSO with RBAC (Admin / User roles).
  Falls back to email header auth during dual-auth window.
  localhost bypass for development.
"""

import asyncio
import base64
import hashlib
import logging
import os
import tempfile
import time
import uuid
from pathlib import Path
from typing import Optional

import pandas as pd
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles

from backend.auth_middleware import (
    AuthenticatedUser,
    require_admin,
    validate_allowed_user,
)
from backend.processor import (
    STANDARD_DF,
    process_schedule,
    process_schedule_with_ai_extraction,
    save_p6_xlsx,
)
from backend.ai_client import clear_extraction_logs, extract_tables_from_pdf, get_extraction_logs
from backend.knowledge_base import get_corrections, get_kb_metadata, list_projects, save_corrections
from backend.risk_knowledge_base import (
    get_risk_corrections,
    get_risk_kb_metadata,
    save_risk_corrections,
)
from backend.activity_store import (
    capture_event,
    complete_event,
    get_event as _get_activity_event,
    list_events,
    log_admin_view,
    EventFilter,
)
from backend.activity_store import init_db as _init_activity_db

# Boot activity DB on import
_init_activity_db()

load_dotenv()

logger = logging.getLogger(__name__)

# ─── App Setup ───────────────────────────────────────────────────────────────

app = FastAPI(
    title="LAP Platform",
    description=(
        "Lease Agentic Planning platform. "
        "IMS Generator: P6 schedule PDF → P6-import XLSX. "
        "Schedule Risk Manager: Assessment Excel → Acumen Fuse risk register."
    ),
    version="2.0.0",
)

FRONTEND_DIR = Path(__file__).parent.parent / "frontend"
app.mount("/app", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="static")

allowed_origins_env = os.getenv(
    "ALLOWED_ORIGINS",
    "https://p6-converter.sta.microsoft.com",
)
allowed_origins = [o.strip() for o in allowed_origins_env.split(",")]
if "http://localhost:8000" not in allowed_origins:
    allowed_origins.append("http://localhost:8000")

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Auth dependency ─────────────────────────────────────────────────────────

async def current_user(request: Request) -> AuthenticatedUser:
    return await validate_allowed_user(request)


# ─── Activity helpers ─────────────────────────────────────────────────────────

def _file_meta(name: str, data: bytes) -> dict:
    """Return a SHA-256 + size descriptor for an uploaded file."""
    return {
        "filename": name,
        "size": len(data),
        "sha256": hashlib.sha256(data).hexdigest(),
    }


def _safe_int(val, default=0) -> int:
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.get("/", tags=["Health"])
async def root():
    return RedirectResponse(url="/app")


@app.get("/health", tags=["Health"])
async def health():
    return {"status": "healthy"}


# ─── Auth ─────────────────────────────────────────────────────────────────────

@app.get("/auth/config", tags=["Auth"])
async def auth_config():
    return JSONResponse({
        "msalClientId": os.getenv("ENTRA_CLIENT_ID", ""),
        "msalTenantId": os.getenv("ENTRA_TENANT_ID", ""),
        "ssoEnabled": bool(os.getenv("ENTRA_TENANT_ID") and os.getenv("ENTRA_CLIENT_ID")),
    })

@app.post("/auth/login", tags=["Auth"])
async def auth_login(request: Request):
    from backend.auth_middleware import ALLOWED_EMAILS

    body = await request.json()
    email = (body.get("email") or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Email is required.")
    if ALLOWED_EMAILS and email not in ALLOWED_EMAILS:
        raise HTTPException(
            status_code=403,
            detail="Your email is not authorised to use this application.",
        )
    name = email.split("@")[0].replace(".", " ").title()
    return {"user": {"email": email, "name": name}}


@app.get("/auth/me", tags=["Auth"])
async def auth_me(request: Request):
    user = await validate_allowed_user(request)
    return {
        "email": user.email,
        "name": user.name,
        "roles": user.roles,
        "is_admin": user.is_admin,
        "source": user.source,
    }


# ─── IMS Generator ───────────────────────────────────────────────────────────

@app.post("/risk-manager/parse/", tags=["Risk Manager"])
async def risk_manager_parse(
    request: Request,
    assessment_file: UploadFile = File(...),
    project_code: str = Form(...),
    tranches: str = Form(...),
    assessment_date: str = Form(...),
    risk_register_file: Optional[UploadFile] = File(default=None),
):
    """Parse schedule assessment Excel and return AI-categorised risk candidates."""
    user = await current_user(request)
    from backend.risk_parser import parse_assessment

    t0 = time.time()
    event_id = capture_event(
        user_email=user.email,
        function_name="schedule_risk_manager",
        project_code=project_code,
        tranche=tranches,
        input_files=[_file_meta(assessment_file.filename or "assessment.xlsx", await assessment_file.read())],
        user_display_name=user.name,
    )
    # Reset file position after reading for meta
    await assessment_file.seek(0)
    file_bytes = await assessment_file.read()
    register_bytes = await risk_register_file.read() if risk_register_file else None

    kb_corrections = await asyncio.to_thread(get_risk_corrections, project_code)
    try:
        result = await asyncio.to_thread(
            parse_assessment, file_bytes, project_code, tranches, assessment_date,
            register_bytes, kb_corrections,
        )
    except Exception as e:
        complete_event(
            event_id=event_id,
            status="failed",
            error_message=str(e),
            duration_ms=int((time.time() - t0) * 1000),
        )
        raise HTTPException(status_code=500, detail=f"Parsing failed: {str(e)}")

    duration_ms = int((time.time() - t0) * 1000)
    candidates = result["risks"]
    complete_event(
        event_id=event_id,
        status="success",
        result_summary={
            "risk_count": len(candidates),
            "manual_scoring_required": result.get("manual_scoring_required", False),
        },
        duration_ms=duration_ms,
    )
    return JSONResponse({
        "project_code": project_code.upper(),
        "tranches": tranches,
        "assessment_date": assessment_date,
        "risk_count": len(candidates),
        "risks": candidates,
        "manual_scoring_required": result["manual_scoring_required"],
    })


@app.post("/convert/", tags=["Convert"])
async def convert(
    request: Request,
    schedule_file: UploadFile = File(...),
    project_code: str = Form(...),
    tranche: str = Form(default="01"),
    data_date: str = Form(default=None),
):
    _ = await current_user(request)
    schedule_name = schedule_file.filename or ""
    if not schedule_name.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="schedule_file must be a PDF (.pdf)")
    try:
        pdf_bytes = await schedule_file.read()
        tables = extract_tables_from_pdf(pdf_bytes)
        if not tables:
            raise HTTPException(status_code=422, detail="No tables found in the PDF.")
        raw_df = max(tables, key=lambda t: t.shape[0] * t.shape[1])
        result_df, _, _ = process_schedule_with_ai_extraction(
            pdf_bytes=pdf_bytes,
            standard_df=STANDARD_DF,
            project_code=project_code,
            tranche=tranche,
            data_date=data_date,
        )
        if result_df.empty:
            result_df = process_schedule(
                raw_df=raw_df,
                standard_df=STANDARD_DF,
                project_code=project_code,
                tranche=tranche,
                data_date=data_date,
            )
        output_path = os.path.join(tempfile.gettempdir(), "p6_import.xlsx")
        save_p6_xlsx(result_df, str(output_path))
        return FileResponse(
            path=str(output_path),
            filename="p6_import.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except HTTPException:
        raise
    except Exception as exc:
        import traceback
        raise HTTPException(
            status_code=500,
            detail=f"Processing failed: {str(exc)}\n\n{traceback.format_exc()}",
        )


# ─── Observability ────────────────────────────────────────────────────────────

@app.get("/extraction-logs/", tags=["Observability"])
async def get_logs():
    logs = get_extraction_logs()
    return JSONResponse({"total_logs": len(logs), "logs": logs})


@app.post("/extraction-logs/clear/", tags=["Observability"])
async def clear_logs():
    clear_extraction_logs()
    return JSONResponse({"message": "All extraction logs cleared"})


@app.post("/debug-extract/", tags=["Debug"])
async def debug_extract(
    request: Request,
    schedule_file: UploadFile = File(...),
    project_code: str = Form(...),
    tranche: str = Form(default="01"),
    data_date: str = Form(default=None),
):
    """Debug endpoint — returns extracted and processed rows."""
    _ = await current_user(request)
    if not (schedule_file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Must be a PDF")
    try:
        pdf_bytes = await schedule_file.read()
        result_df, all_activities, _ = process_schedule_with_ai_extraction(
            pdf_bytes=pdf_bytes,
            standard_df=STANDARD_DF,
            project_code=project_code,
            tranche=tranche,
            data_date=data_date,
        )
        logs = get_extraction_logs()
        pdf_log = next(
            (l for l in reversed(logs) if l.get("event_type") == "PDF_EXTRACTION_START"),
            {},
        )
        log_details = pdf_log.get("details", {})
        matched = [a for a in all_activities if a.get("matched_std_id")]
        return JSONResponse({
            "debug_info": {
                "pdf_filename": schedule_file.filename or "unknown.pdf",
                "pdf_text_length": log_details.get("pdf_text_length", 0),
                "pages_returned": log_details.get("pages_returned", 0),
                "tables_found": log_details.get("tables_found", 0),
                "ai_deployment": os.getenv("AZURE_OPENAI_DEPLOYMENT", ""),
            },
            "extraction_summary": {
                "total_extracted": len(all_activities),
                "total_matched": len(matched),
                "total_output": len(result_df),
            },
            "activity_details": all_activities,
            "processed_activities": result_df.to_dict(orient="records"),
        })
    except HTTPException:
        raise
    except Exception as exc:
        import traceback
        return JSONResponse({"error": str(exc), "traceback": traceback.format_exc()}, status_code=500)


# ─── Batch helpers ────────────────────────────────────────────────────────────

async def _process_project(
    pdf_bytes: bytes,
    pdf_filename: str,
    project_code: str,
    tranche: str,
    data_date: Optional[str],
    pages: Optional[str] = None,
) -> dict:
    kb = await asyncio.to_thread(get_corrections, project_code)
    result_df, _, frontend_rows = await asyncio.to_thread(
        process_schedule_with_ai_extraction,
        pdf_bytes, STANDARD_DF, project_code, tranche, data_date, pages, kb,
    )
    tmp_path = os.path.join(tempfile.gettempdir(), f"p6_{project_code}_{uuid.uuid4().hex[:6]}.xlsx")
    save_p6_xlsx(result_df, tmp_path)
    with open(tmp_path, "rb") as f:
        xlsx_b64 = base64.b64encode(f.read()).decode("utf-8")
    os.unlink(tmp_path)
    kb_meta = await asyncio.to_thread(get_kb_metadata, project_code)
    return {
        "project_code": project_code,
        "filename": pdf_filename,
        "row_count": len(result_df),
        "activities": frontend_rows,
        "xlsx_b64": xlsx_b64,
        "kb_corrections_used": len(kb),
        "kb_last_updated": (kb_meta or {}).get("last_updated"),
    }


# ─── Protected IMS Endpoints ─────────────────────────────────────────────────

@app.post("/batch-convert/", tags=["Convert"])
async def batch_convert(request: Request, **kwargs):
    """
    Convert up to 3 PDF schedules in parallel.
    Requires authentication. Instruments each project slot as an activity event.
    """
    # Accept all Form fields dynamically so we don't repeat the long signature
    user = await current_user(request)
    logger.info("batch-convert by %s", user.email)

    form_data = await request.form()
    tasks, task_meta = [], []
    for slot in range(1, 4):
        file_key = f"schedule_file_{slot}"
        code_key = f"project_code_{slot}"
        tranche_key = f"tranche_{slot}"
        date_key = f"data_date_{slot}"
        pages_key = f"pages_{slot}"

        file: UploadFile = form_data.get(file_key)
        code: str = form_data.get(code_key, "")
        tranche_val = form_data.get(tranche_key, "01") or "01"
        date_val = form_data.get(date_key)
        pages_val = form_data.get(pages_key)

        if file is None or not code:
            continue
        if not (file.filename or "").lower().endswith(".pdf"):
            raise HTTPException(
                status_code=400,
                detail=f"File for project '{code}' must be a PDF.",
            )
        pdf_bytes = await file.read()
        tasks.append(
            _process_project(pdf_bytes, file.filename or "schedule.pdf",
                             code, tranche_val, date_val, pages_val)
        )
        task_meta.append({
            "code": code, "tranche": tranche_val,
            "filename": file.filename, "bytes": len(pdf_bytes),
        })

    if not tasks:
        raise HTTPException(
            status_code=400,
            detail="At least one PDF file with a project code is required.",
        )

    # Capture parent event for the aggregate batch
    t0 = time.time()
    input_files = [{"filename": m["filename"], "size": m["bytes"],
                    "sha256": hashlib.sha256(b"\0").hexdigest()} for m in task_meta]
    event_id = capture_event(
        user_email=user.email,
        function_name="ims_generator",
        project_code=", ".join(m["code"] for m in task_meta),
        input_files=input_files,
        user_display_name=user.name,
    )

    results = await asyncio.gather(*tasks, return_exceptions=True)
    duration_ms = int((time.time() - t0) * 1000)

    projects = []
    errors, successes = [], []
    for i, r in enumerate(results):
        if isinstance(r, Exception):
            import traceback as tb
            projects.append({"error": str(r), "traceback": tb.format_exc()})
            errors.append(task_meta[i]["code"])
        else:
            projects.append(r)
            successes.append(task_meta[i]["code"])

    # Complete parent event
    status = "failed" if errors and not successes else "partial" if errors else "success"
    complete_event(
        event_id=event_id,
        status=status,
        result_summary={
            "projects_attempted": len(task_meta),
            "projects_succeeded": len(successes),
            "projects_failed": len(errors),
            "error_projects": errors,
        },
        duration_ms=duration_ms,
    )
    return JSONResponse({"projects": projects})


@app.post("/download-edited/", tags=["Convert"])
async def download_edited(request: Request):
    """Generate XLSX from user-edited activities."""
    _ = await current_user(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    project_code = body.get("project_code", "unknown")
    activities = body.get("activities", [])
    if not activities:
        raise HTTPException(status_code=400, detail="No activities provided")

    from backend.processor import P6_COLUMNS, P6_HEADERS
    from io import BytesIO

    clean = [{k: v for k, v in a.items() if k != "_idx"} for a in activities]
    df = pd.DataFrame(clean, columns=P6_COLUMNS)
    for col in ["cstr_date", "act_start_date", "act_end_date"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], format="%m/%d/%Y %H:%M", errors="coerce")
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame([P6_COLUMNS]).to_excel(writer, sheet_name="TASK", index=False, header=False, startrow=0)
        pd.DataFrame([P6_HEADERS]).to_excel(writer, sheet_name="TASK", index=False, header=False, startrow=1)
        df.to_excel(writer, sheet_name="TASK", index=False, header=False, startrow=2)
    buffer.seek(0)
    xlsx_b64 = base64.b64encode(buffer.read()).decode("utf-8")
    return JSONResponse({"xlsx_b64": xlsx_b64})


# ─── Risk Manager ──────────────────────────────────────────────────────────────

@app.post("/risk-manager/kb/save/", tags=["Risk Manager"])
async def risk_kb_save(request: Request):
    """Save reviewed risk candidates as the Knowledge Base."""
    _ = await current_user(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    project_code = (body.get("project_code") or "").strip().upper()
    risks = body.get("risks", [])
    if not project_code:
        raise HTTPException(status_code=400, detail="project_code is required")
    if not risks:
        raise HTTPException(status_code=400, detail="risks list is empty")
    ok = await asyncio.to_thread(save_risk_corrections, project_code, risks)
    if not ok:
        raise HTTPException(status_code=503, detail="Risk KB unavailable")
    meta = await asyncio.to_thread(get_risk_kb_metadata, project_code)
    return JSONResponse({
        "message": f"Risk knowledge base saved for {project_code}",
        "project_code": project_code,
        "corrections_saved": (meta or {}).get("count", 0),
        "last_updated": (meta or {}).get("last_updated"),
    })


_SCORE_LABEL = {1: "Negligible", 2: "Very Low", 3: "Low", 4: "Medium", 5: "High", 6: "Very High"}

@app.post("/risk-manager/download/", tags=["Risk Manager"])
async def risk_manager_download(request: Request):
    """Generate Acumen Fuse-compatible risk register XLSX."""
    _ = await current_user(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    project_code = (body.get("project_code") or "RISK").strip().upper()
    risks = [r for r in body.get("risks", []) if not r.get("_removed")]
    if not risks:
        raise HTTPException(status_code=400, detail="No risks to export")

    xlsx_bytes = await asyncio.to_thread(_build_acumen_xlsx, project_code, risks)
    return JSONResponse({"xlsx_b64": base64.b64encode(xlsx_bytes).decode()})


def _build_acumen_xlsx(project_code: str, risks: list[dict]) -> bytes:
    """Generate an Acumen Fuse-compatible risk register XLSX."""
    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADERS = [
        "Enabled", "Absolute M", "ID", "Type", "Name", "Shape",
        "Current Probability", "Current Schedule", "Current Cost", "Current Score",
        "Mitigation Enabled", "Mitigation Description", "Mitigation Duration", "Mitigation Cost",
        "Mitigated Probability", "Mitigated Schedule", "Mitigated Cost", "Mitigated Score",
        "Calendar", "No Split",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    navy_fill = PatternFill("solid", fgColor="1A3C6E")
    white_bold = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    row_num = 2
    for risk_idx, r in enumerate(risks, 1):
        prob = r.get("current_probability", 0)
        sched = r.get("current_schedule", 0)
        mit_p = r.get("mitigated_probability", "")
        mit_s = r.get("mitigated_schedule", "")
        try:
            score = int(prob) * int(sched)
        except (TypeError, ValueError):
            score = ""

        mit_desc = r.get("mitigation_description", "") or ""
        row_data = [
            "Yes", "No", f"R{risk_idx}", r.get("type", "Threat"), r.get("risk_name", ""),
            r.get("shape", "Triangular"),
            _SCORE_LABEL.get(int(prob), ""),
            _SCORE_LABEL.get(int(sched), ""),
            "Negligible",
            score,
            "Yes" if mit_desc.strip() else "No",
            mit_desc,
            r.get("mitigation_duration", "") or "",
            r.get("mitigation_cost", "") or "",
            _SCORE_LABEL.get(int(mit_p)) if mit_p else "",
            _SCORE_LABEL.get(int(mit_s)) if mit_s else "",
            "", "", "", "No",
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row_num += 1

    for col_idx, header in enumerate(HEADERS, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, max_row=row_num - 1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value) if cell.value else ""))
                except Exception:
                    pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
    ws.row_dimensions[1].height = 30

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Knowledge Base ───────────────────────────────────────────────────────────

@app.post("/knowledge-base/save/", tags=["Knowledge Base"])
async def kb_save(request: Request):
    """Save planner-corrected activities as the Knowledge Base."""
    _ = await current_user(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")
    project_code = (body.get("project_code") or "").strip().upper()
    activities = body.get("activities", [])
    if not project_code:
        raise HTTPException(status_code=400, detail="project_code is required")
    if not activities:
        raise HTTPException(status_code=400, detail="activities list is empty")
    ok = await asyncio.to_thread(save_corrections, project_code, activities)
    if not ok:
        raise HTTPException(status_code=503, detail="KB unavailable")
    meta = await asyncio.to_thread(get_kb_metadata, project_code)
    return JSONResponse({
        "message": f"Knowledge base saved for {project_code}",
        "project_code": project_code,
        "corrections_saved": (meta or {}).get("count", 0),
        "last_updated": (meta or {}).get("last_updated"),
    })


@app.get("/knowledge-base/", tags=["Knowledge Base"])
async def kb_list():
    """List all projects with a Knowledge Base entry."""
    projects = await asyncio.to_thread(list_projects)
    results = []
    for code in projects:
        meta = await asyncio.to_thread(get_kb_metadata, code)
        results.append({
            "project_code": code,
            "corrections": (meta or {}).get("count", 0),
            "last_updated": (meta or {}).get("last_updated"),
        })
    return JSONResponse({"projects": results})


@app.delete("/knowledge-base/{project_code}/", tags=["Knowledge Base"])
async def kb_delete(project_code: str):
    """Delete the Knowledge Base entry for a project."""
    from backend.knowledge_base import _get_client, _CONTAINER

    client = await asyncio.to_thread(_get_client)
    if not client:
        raise HTTPException(status_code=503, detail="KB unavailable")
    try:
        blob = client.get_blob_client(_CONTAINER, f"{project_code.upper()}.json")
        await asyncio.to_thread(blob.delete_blob)
    except Exception:
        raise HTTPException(status_code=404, detail=f"No KB entry for {project_code}")
    return JSONResponse({"message": f"KB deleted for {project_code.upper()}"})


# ─── Admin ────────────────────────────────────────────────────────────────────

@app.get("/api/admin/events", tags=["Admin"])
async def admin_list_events(request: Request):
    """
    List activity events with server-side filter, sort, and pagination.
    Admin role required.
    """
    admin = await require_admin(request)

    page = _safe_int(request.query_params.get("page", 1), 1)
    limit = min(_safe_int(request.query_params.get("limit", 50), 50), 200)
    filters = EventFilter(
        user_email=request.query_params.get("user_email"),
        function_name=request.query_params.get("function_name"),
        date_from=request.query_params.get("date_from"),
        date_to=request.query_params.get("date_to"),
        status=request.query_params.get("status"),
        project_code=request.query_params.get("project_code"),
        page=page,
        limit=limit,
        sort_col=request.query_params.get("sort_col", "triggered_at"),
        sort_dir=request.query_params.get("sort_dir", "desc"),
    )
    result = list_events(filters)
    return JSONResponse(result)


@app.get("/api/admin/events/{event_id}", tags=["Admin"])
async def admin_get_event(request: Request, event_id: str):
    """
    Get a single activity event by ID.
    Admin role required. Logs the view to admin_views for self-audit.
    """
    admin = await require_admin(request)

    ev = _get_activity_event(event_id)
    if not ev:
        raise HTTPException(status_code=404, detail="Event not found")

    # Self-audit: log that this admin viewed this event
    log_admin_view(
        admin_email=admin.email,
        admin_display_name=admin.name,
        target_event_id=event_id,
        target_function=ev.get("function_name", ""),
        target_project=ev.get("project_code"),
    )

    return JSONResponse(ev)


# ─── Run locally ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn

    port = int(os.getenv("PORT", "8000"))
    uvicorn.run("backend.main:app", host="0.0.0.0", port=port, reload=True)
